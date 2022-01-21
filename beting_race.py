import re, requests, os, time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--user-data-dir=C:\\Users\\Angus\\AppData\\Local\\Google\\Chrome\\Selenium")

EXCEL = {
    'calculator': 'DOUBLES CALCULATOR 1.xlsx',
    'summary': 'Stats for Doubles.xlsx'
}

RACE_API = 'https://www.sportsbet.com.au/apigw/sportsbook-racing/Sportsbook/Racing/Events/{}/RacecardWithContext'
REGEX = 'https://www.sportsbet.com.au/horse-racing/([a-zA-Z0-9-]+)/([a-zA-Z0-9-]+)/race-([0-9]+)-([0-9]+)/running-double-([0-9]+)'

url = input("Enter url: ")
tab = float(input("TAB Winnings: "))
win = float(input("What is the minimum profit you will accept? "))

matches = re.search(REGEX, url)
print("dddddddddddddddddddddddddddddddddd", matches)

# identify race information using URL
country, competition, race_number, event_id, pool_id = matches.groups()
# get race data
data = requests.get(RACE_API.format(event_id)).json()

# find the two legs 
pool = list(filter(lambda x: x['id'] == int(pool_id), data['racecardContext']['pools']))[0]
legs = pool['eventIds']
date = datetime.fromtimestamp(data['racecardEvent']['competitionStartTime']).strftime('%d-%b-%Y')

race_numbers = []
legs_data = []
distance = []

for leg in legs:
    # get data specific to a leg
    data = requests.get(RACE_API.format(leg)).json()['racecardEvent']
    # get race number of that leg
    race_numbers.append(data['raceNumber'])
    # get track distance for each race
    distance.append(data['distance'])
    # remove horses not in the race
    horses = list(filter(lambda horse: horse['statusCode'] == 'A', data['markets'][0]['selections']))

    # remember which horse is in the ith position in the list before sorting
    for i in range(len(horses)):
        horses[i]['index'] = i
    # sort horses based on win percentage
    sorted_horses = sorted(list(map(lambda horse: (
        horse['statistics']['winPercentage'] / 100,
        horse['runnerNumber'],
        horse['prices'][0]['winPrice'],
        horse['index'],
        horse['statistics']['total'].split(':')[0]), horses)), reverse = True)
    # only get 
    legs_data.append(sorted_horses)

print(f"The first leg has {len(legs_data[0])} horses.")
print(f"The second leg has {len(legs_data[1])} horses.")

while True:
    size = list(map(int, input("Enter bet type (i.e MxN): ").split('x')))
    if size[0] <= 0:
        print('Please enter positive number of horses')
        continue
    if size[1] <= 0:
        print('Please enter positive number of horses')
        continue
    if size[0] > len(legs_data[0]):
        print('Too many horses on the first leg.')
        continue
    if size[1] > len(legs_data[1]): 
        print('Too many horses on the second leg.')
        continue
    break

output_filename = f'computations/{competition}-R{race_numbers[0]}&R{race_numbers[1]}'
print(f'Saving results into {output_filename}.xlsx')

workbook = load_workbook(filename = EXCEL['calculator'])
worksheet = workbook['DOUBLES']
worksheet['D5'] = tab
for col in range(2):
    for i in range(size[col]):
        worksheet['JL'[col] + str(i + 1)] = legs_data[col][i][1]
        worksheet['KM'[col] + str(i + 1)] = legs_data[col][i][0]
        worksheet['BD'[col] + str(9 + i)] = legs_data[col][i][2]

worksheet['D1'] = competition
worksheet['D2'] = date
worksheet['D3'] = str(race_numbers[0]) + ',' + str(race_numbers[1])
workbook.save(filename = f'{output_filename}.xlsx')

def bet(x, y, tab):
    if x == 0 or y == 0:
        odds = 0
    else:
        odds = (x + 1) * (y + 1) - 1
    if x == 0 or y == 0:
        outlay = 0
    else:
        outlay = tab / (odds + 1)
    return outlay

bets = []
for i in range(size[0]):
    for j in range(size[1]):
        bets.append((legs_data[0][i][1], legs_data[1][j][1], bet(legs_data[0][i][2], legs_data[1][j][2], float(tab)), legs_data[0][i][3], legs_data[1][j][3]))

outlay = sum(map(lambda x: x[2], bets))

print("-- BET SUMMARY --")
print(f"You will be making {size[0] * size[1]} bets")
print(f"Top {size[0]} horses for first leg:", ', '.join(map(lambda x: str(x[1]), legs_data[0][:size[0]])))
print(f"Top {size[1]} horses for first leg:", ', '.join(map(lambda x: str(x[1]), legs_data[1][:size[1]])))
print("TAB Winnings =", tab)
print("Your total outlay will be", outlay)
print("Your expected profit will be", tab - outlay)
if tab - outlay >= win:
    input("Press enter to proceed betting.")
else:
    input("Insufficient profit to proceed with bets.")
    exit()

driver = webdriver.Chrome('./chromedriver.exe', options=chrome_options)

driver.get(url)
input('test')
try:
    clear = driver.find_element(By.XPATH, "//div[@data-automation-id='betslip-clear-button']")
    input("Press enter to clear bet slips and to start betting.")
    driver.execute_script("arguments[0].children[0].children[1].click();", clear)
except:
    input("Press enter to start betting.")

number_of_bets = 0
for x, y, z, i, j in bets:
    if number_of_bets == 25:
        input('press enter to continue - max bets reached')
    while True:
        elems = driver.find_elements(By.XPATH, "//div[starts-with(@data-automation-id,'racecard-exotic-')]")
        if len(elems) > 0: break    

    first_leg_buttons = []
    for elem in elems[1:]:
        if elem.get_attribute('data-automation-id')[-3:] == 'off':
            first_leg_buttons.append(elem)

    driver.execute_script("arguments[0].click();", first_leg_buttons[i])
    next_leg = driver.find_element(By.XPATH, "//div[@data-automation-id='next-race']")
    driver.execute_script("arguments[0].children[0].click();", next_leg)
    while True:
        elems = driver.find_elements(By.XPATH, "//div[starts-with(@data-automation-id,'racecard-exotic-')]")
        if len(elems) > 0: break    

    second_leg_buttons = []
    for elem in elems[1:]:
        if elem.get_attribute('data-automation-id')[-3:] == 'off':
            second_leg_buttons.append(elem)
    driver.execute_script("arguments[0].click();", second_leg_buttons[j])

    stake = driver.find_element(By.XPATH, "//input[@data-automation-id='exotic-stake-amount']")
    stake.send_keys(Keys.CONTROL + "a")
    stake.send_keys(Keys.DELETE)
    stake.send_keys(str(z))

    stake.send_keys(Keys.RETURN)

    view_selection = driver.find_element(By.XPATH, "//div[@data-automation-id='accordion-header']")
    driver.execute_script("arguments[0].click();", view_selection)

    clear = driver.find_element(By.XPATH, "//button[@data-automation-id='racecard-selections-accordion-clear-button']")
    driver.execute_script("arguments[0].click();", clear)

    clear = driver.find_element(By.XPATH, "//div[@data-automation-id='focus-modal-action-button']")
    driver.execute_script("arguments[0].children[0].children[1].click();", clear)

    previous_leg = driver.find_element(By.XPATH, "//div[@data-automation-id='previous-race']")
    driver.execute_script("arguments[0].children[0].click();", previous_leg)
    number_of_bets += 1

print('The bet has been finished!')

input(f'Press enter to write the results to {EXCEL["summary"]}')

workbook = load_workbook(filename = EXCEL['summary'])

worksheet = workbook['Sheet1']
print(worksheet)

row = 3
while True:
    print(worksheet['A' + str(row)].value)
    if worksheet['A' + str(row)].value in [None, '']:
        break
    row += 1

worksheet['A' + str(row)] = str(datetime.now())
worksheet['B' + str(row)] = country
worksheet['C' + str(row)] = data['weather']
worksheet['D' + str(row)] = str(race_numbers[0])
worksheet['E' + str(row)] = str(len(legs_data[0]))
worksheet['F' + str(row)] = str(sum(map(lambda x: int(x[4]), legs_data[0])) / len(legs_data[0]))
worksheet['G' + str(row)] = distance[0]
worksheet['H' + str(row)] = str(race_numbers[1])
worksheet['I' + str(row)] = str(len(legs_data[1]))
worksheet['J' + str(row)] = str(sum(map(lambda x: int(x[4]), legs_data[1])) / len(legs_data[1]))
worksheet['K' + str(row)] = distance[1]
worksheet['L' + str(row)] = tab
worksheet['M' + str(row)] = win
worksheet['N' + str(row)] = outlay

workbook.save(filename = EXCEL['summary'])
